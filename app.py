import asyncio
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime
import hashlib
import json
import logging
import os
import re
import shutil
import sys
import tempfile
import time
import unicodedata
from urllib.parse import quote, unquote, urlsplit
import calendar

import aiofiles
from fastapi import FastAPI, File, Query, Request, UploadFile
from fastapi.middleware.gzip import GZipMiddleware
from fastapi.responses import HTMLResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook

# Try orjson for faster JSON serialization
try:
    import orjson
    HAS_ORJSON = True

    class ORJSONResponse(Response):
        media_type = "application/json"

        def render(self, content) -> bytes:
            return orjson.dumps(
                content,
                option=orjson.OPT_NON_STR_KEYS | orjson.OPT_SERIALIZE_NUMPY,
            )
    DefaultJSONResponse = ORJSONResponse
except ImportError:
    HAS_ORJSON = False
    DefaultJSONResponse = JSONResponse

logger = logging.getLogger(__name__)

app = FastAPI(default_response_class=DefaultJSONResponse)
app.add_middleware(GZipMiddleware, minimum_size=1024, compresslevel=5)
app.mount("/static", StaticFiles(directory="static"), name="static")

TYPE_MAP = {
    "CESP": "Espèces",
    "CTRT": "Traite",
    "CCHQR": "Chèque",
}

VALID_SITES = {"SFX", "MAH", "NAB", "SSE", "TUN"}

# Pre-compiled regex patterns (huge perf boost)
CAM_REGEX = re.compile(r"(CAM\d+)")
DATE_REGEX = re.compile(r"^\d{8}$")

# ── Upload configuration ───────────────────────────────────────────────────────
MAX_TOTAL_UPLOAD_SIZE = int(os.environ.get("MAX_TOTAL_UPLOAD_SIZE", 500 * 1024 * 1024))
MAX_SINGLE_FILE_SIZE = int(os.environ.get("MAX_SINGLE_FILE_SIZE", 100 * 1024 * 1024))
CHUNK_SIZE = int(os.environ.get("CHUNK_SIZE", 1024 * 1024))  # 1 MB (was 8 KB)
MAX_CONCURRENT_UPLOADS = int(os.environ.get("MAX_CONCURRENT_UPLOADS", 8))
MAX_PARSE_WORKERS = int(os.environ.get("MAX_PARSE_WORKERS", min(8, (os.cpu_count() or 4))))


def format_size(bytes_size: int) -> str:
    if bytes_size < 1024:
        return f"{bytes_size} octets"
    elif bytes_size < 1024 * 1024:
        return f"{bytes_size / 1024:.1f} Ko"
    elif bytes_size < 1024 * 1024 * 1024:
        return f"{bytes_size / (1024 * 1024):.1f} Mo"
    else:
        return f"{bytes_size / (1024 * 1024 * 1024):.2f} Go"


# ── Source configuration ───────────────────────────────────────────────────────
DEFAULT_CURRENT_REGLEMENT_FILE = os.environ.get("CURRENT_REGLEMENT_FILE", "")
DEFAULT_HISTORY_REGLEMENTS_DIR = os.environ.get("HISTORY_REGLEMENTS_DIR", "")
WINDOWS_SYNC_CACHE_DIR = os.environ.get(
    "WINDOWS_SYNC_CACHE_DIR",
    os.path.join(tempfile.gettempdir(), "sind_reglement_app", "reglements_cache"),
)

# ── In-memory data cache ──────────────────────────────────────────────────────
_cache: dict = {
    "all_rows": [],
    "current_rows": [],
    "source_files": [],
    "current_source_files": [],
    "article_lookup": {},
    "etatmarge_lookup": {},
    "etatmarge_warnings": [],
    "all_facture_lines": [],
    "all_big_factures": [],
    "current_big_factures": [],
    "facture_source_files": [],
    "current_facture_source_files": [],
    "client_lookup": {},
    "clients_by_cam": {},
    "client_warnings": [],
    "warnings": [],
    "current_warnings": [],
    "loaded_at": None,
    "coverage_start": None,
    "coverage_end": None,
    "history_file_count": 0,
    "facture_coverage_start": None,
    "facture_coverage_end": None,
    "facture_history_file_count": 0,
    "source_diagnostics": {},
    "sync": {},
    "import_context": {},
    # Precomputed payloads cache
    "_default_payload": None,
    "_status_payload": None,
    "_range_cache": {},  # key: (start, end) -> payload
}

# Larger thread pool for parallel I/O and parsing
IMPORT_EXECUTOR = ThreadPoolExecutor(max_workers=MAX_PARSE_WORKERS, thread_name_prefix="parse")
_import_tasks: set[asyncio.Task] = set()

CAM_SITE_MAP: dict[str, str] = {
    "CAM01": "SFX", "CAM02": "SFX", "CAM03": "SFX", "CAM04": "SFX",
    "CAM05": "SFX", "CAM06": "SFX", "CAM07": "SFX", "CAM36": "SFX",
    "CAM37": "SFX", "CAM38": "SFX", "CAM48": "SFX", "CAM49": "SFX",
    "CAM58": "SFX", "CAM59": "SFX",
    "CAM40": "MAH", "CAM41": "MAH", "CAM42": "MAH", "CAM43": "MAH",
    "CAM44": "MAH", "CAM45": "MAH", "CAM57": "MAH",
    "CAM50": "NAB", "CAM51": "NAB", "CAM52": "NAB", "CAM53": "NAB",
    "CAM54": "NAB",
    "CAM08": "SSE", "CAM09": "SSE", "CAM10": "SSE", "CAM11": "SSE",
    "CAM12": "SSE", "CAM13": "SSE", "CAM14": "SSE", "CAM15": "SSE",
    "CAM39": "SSE", "CAM46": "SSE", "CAM47": "SSE",
    "CAM16": "TUN", "CAM17": "TUN", "CAM18": "TUN", "CAM19": "TUN",
    "CAM20": "TUN", "CAM21": "TUN", "CAM22": "TUN", "CAM23": "TUN",
    "CAM24": "TUN", "CAM25": "TUN", "CAM26": "TUN", "CAM27": "TUN",
    "CAM29": "TUN", "CAM30": "TUN", "CAM31": "TUN",
}


def get_site(cam: str | None) -> str:
    if cam is None:
        return "Inconnu"
    return CAM_SITE_MAP.get(cam, "Inconnu")


def normalize_site(site: str | None) -> str:
    if site is None:
        return "Inconnu"
    site = site.strip().upper()
    return site if site in VALID_SITES else "Inconnu"


def is_file_uri(source: str) -> bool:
    return bool(source and source.startswith("file://"))

def normalize_token(value: str) -> str:
    """Aggressive normalization: removes accents, spaces, underscores, dashes, converts to lowercase."""
    if not value:
        return ""
    # Remove accents (é → e, ç → c, etc.)
    normalized = unicodedata.normalize("NFKD", value)
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    # Convert to lowercase
    normalized = normalized.casefold()
    # Remove all separators (spaces, underscores, dashes, dots)
    normalized = re.sub(r"[\s_\-\.]+", "", normalized)
    return normalized

def last_day_of_month(year: int, month: int) -> int:
    return calendar.monthrange(year, month)[1]

def shift_date_back_one_month(d: date) -> date:
    year, month = d.year, d.month - 1
    if month == 0:
        month = 12
        year -= 1
    day = min(d.day, last_day_of_month(year, month))
    return date(year, month, day)

def shift_date_back_one_year(d: date) -> date:
    year = d.year - 1
    month, day = d.month, d.day
    if month == 2 and day == 29 and not calendar.isleap(year):
        day = 28
    return date(year, month, day)

def is_full_calendar_month(start: date, end: date) -> bool:
    return (
        start.day == 1
        and start.year == end.year and start.month == end.month
        and end.day == last_day_of_month(end.year, end.month)
    )

def is_full_calendar_year(start: date, end: date) -> bool:
    return (
        start.month == 1 and start.day == 1
        and end.month == 12 and end.day == 31
        and start.year == end.year
    )

def compute_previous_period(start: date, end: date) -> tuple[date, date]:
    """
    - Full calendar month  -> previous full calendar month (respects real day counts).
    - Full calendar year   -> previous full calendar year.
    - Anything else (partial month, custom range, month-to-date) ->
      same range shifted back exactly one calendar month, day-clamped.
    """
    if is_full_calendar_year(start, end):
        return shift_date_back_one_year(start), shift_date_back_one_year(end)
    if is_full_calendar_month(start, end):
        prev_start = shift_date_back_one_month(start)
        prev_end = date(prev_start.year, prev_start.month, last_day_of_month(prev_start.year, prev_start.month))
        return prev_start, prev_end
    return shift_date_back_one_month(start), shift_date_back_one_month(end)

def compute_period_totals_scoped(rows, factures, start, end, *, cam=None, site=None):
    def row_match(r):
        if cam and r.get("cam") != cam:
            return False
        if site:
            r_site = get_site(r.get("cam")) if r.get("cam") else r.get("site")
            if r_site != site:
                return False
        return True

    def fac_match(f):
        if cam and f.get("cam") != cam:
            return False
        if site and f.get("site") != site:
            return False
        return True

    filtered_rows = [r for r in filter_rows_by_date(rows, start, end) if row_match(r)]
    filtered_factures = [f for f in filter_big_factures_by_date(factures, start, end) if fac_match(f)]
    return {
        "total_reglements": round(sum(r["amount"] for r in filtered_rows), 3),
        "nb_operations": len(filtered_rows),
        "nb_factures": len(filtered_factures),
        "total_ventes": round(sum(f["total_amount"] for f in filtered_factures), 3),
    }

def build_trend(current: float, previous: float) -> dict:
    delta = round(current - previous, 3)
    pct = None if not previous else round((delta / previous) * 100, 1)
    trend = "up" if delta > 0.0005 else "down" if delta < -0.0005 else "flat"
    return {"current": current, "previous": previous, "delta": delta, "pct": pct, "trend": trend}

def build_comparison_payload_scoped(cache, start, end, *, cam=None, site=None):
    prev_start, prev_end = compute_previous_period(start, end)
    current = compute_period_totals_scoped(cache["all_rows"], cache["all_big_factures"], start, end, cam=cam, site=site)
    previous = compute_period_totals_scoped(cache["all_rows"], cache["all_big_factures"], prev_start, prev_end, cam=cam, site=site)
    return {
        "current_period": {"start": start.isoformat(), "end": end.isoformat()},
        "previous_period": {"start": prev_start.isoformat(), "end": prev_end.isoformat()},
        "reglements": build_trend(current["total_reglements"], previous["total_reglements"]),
        "operations": build_trend(current["nb_operations"], previous["nb_operations"]),
        "factures": build_trend(current["nb_factures"], previous["nb_factures"]),
        "ventes": build_trend(current["total_ventes"], previous["total_ventes"]),
    }

def is_reglements_dir_name(name: str) -> bool:
    return normalize_token(name) == "reglements"


def is_reglement_text_filename(name: str) -> bool:
    base, ext = os.path.splitext(name or "")
    return ext.casefold() == ".txt" and "reglement" in normalize_token(base)


def is_factures_dir_name(name: str) -> bool:
    return normalize_token(name) == "factures"


def is_named_source_file(name: str, expected_name: str) -> bool:
    base, ext = os.path.splitext(name or "")
    normalized = normalize_token(base or name)
    return normalized == normalize_token(expected_name) and ext.casefold() in {"", ".txt", ".csv"}


def is_named_excel_source_file(name: str, expected_name: str) -> bool:
    """
    Match file name against expected name with aggressive normalization.
    
    Examples that should ALL match "etatmarge":
      - EtatMarge.xlsx
      - etatmarge.xlsx
      - Etat Marge.xlsx
      - etat_marge.xlsx
      - état marge.xlsx
      - ETATMARGE.xlsm
    """
    if not name:
        return False
    
    # Split extension
    base, ext = os.path.splitext(name)
    
    # Normalize both the filename and expected name
    normalized_base = normalize_token(base)
    normalized_expected = normalize_token(expected_name)
    
    # Check extension is valid
    valid_extensions = {"", ".txt", ".csv", ".xlsx", ".xlsm", ".xls"}
    
    return (
        normalized_base == normalized_expected 
        and ext.casefold() in valid_extensions
    )

def normalize_article_code(value: str | None) -> str:
    return (value or "").strip().upper()


def discover_uploaded_sources(root_dir: str) -> dict:
    current_file: str | None = None
    history_dir: str | None = None
    article_file: str | None = None
    etatmarge_file: str | None = None
    current_facture_file: str | None = None
    factures_dir: str | None = None
    client_file: str | None = None

    try:
        root_entries = sorted(os.listdir(root_dir), key=str.casefold)
    except OSError:
        root_entries = []

    for name in root_entries:
        candidate = os.path.join(root_dir, name)
        if not current_file and os.path.isfile(candidate) and name.casefold() == "reglement.txt":
            current_file = candidate
        if not history_dir and os.path.isdir(candidate) and is_reglements_dir_name(name):
            history_dir = candidate
        if not article_file and os.path.isfile(candidate) and is_named_source_file(name, "ARTICLE"):
            article_file = candidate
        if not etatmarge_file and os.path.isfile(candidate) and is_named_excel_source_file(name, "etatmarge"):
            etatmarge_file = candidate
        if not current_facture_file and os.path.isfile(candidate) and is_named_source_file(name, "FACTURE"):
            current_facture_file = candidate
        if not factures_dir and os.path.isdir(candidate) and is_factures_dir_name(name):
            factures_dir = candidate
        if not client_file and os.path.isfile(candidate) and is_named_source_file(name, "CLIENT"):
            client_file = candidate
        if (current_file and history_dir and article_file and current_facture_file
                and factures_dir and etatmarge_file and client_file):
            break

    warnings: list[str] = []
    if not current_file:
        warnings.append("Le fichier attendu REGLEMENT.txt est introuvable dans le dossier importé.")
    if not history_dir:
        warnings.append("Le dossier attendu Réglements est introuvable dans le dossier importé.")

    return {
        "root_path": root_dir,
        "root_name": os.path.basename(os.path.normpath(root_dir)) or "Fichiers Sources",
        "current_path": current_file or "",
        "history_path": history_dir or "",
        "article_path": article_file or "",
        "etatmarge_path": etatmarge_file or "",
        "facture_path": current_facture_file or "",
        "factures_path": factures_dir or "",
        "client_path": client_file or "",
        "current_found": bool(current_file),
        "history_found": bool(history_dir),
        "article_found": bool(article_file),
        "etatmarge_found": bool(etatmarge_file),
        "facture_found": bool(current_facture_file),
        "factures_found": bool(factures_dir),
        "client_found": bool(client_file),
        "warnings": warnings,
    }


def get_file_uri_mount_root() -> str | None:
    mount_root = os.environ.get("FILE_URI_MOUNT_ROOT", "").strip()
    return mount_root or None


def get_windows_sync_cache_dir() -> str:
    return os.path.abspath(WINDOWS_SYNC_CACHE_DIR)


def resolve_source_path(source: str) -> tuple[str, dict]:
    if not is_file_uri(source):
        return source, {"path_strategy": "raw_path"}

    parsed = urlsplit(source)
    path_part = unquote(parsed.path or "")
    host = parsed.netloc
    is_remote = bool(host and host.lower() not in {"localhost", "127.0.0.1", "::1"})
    mount_root = get_file_uri_mount_root()

    if is_remote:
        if os.name == "nt":
            resolved = f"\\\\{host}{path_part.replace('/', chr(92))}".rstrip("\\")
            return resolved, {"path_strategy": "windows_unc", "uri_host": host}
        if mount_root:
            segments = [seg for seg in path_part.split("/") if seg and seg not in {".", ".."}]
            mount_root_abs = os.path.abspath(mount_root)
            resolved_candidate = os.path.abspath(os.path.normpath(os.path.join(mount_root_abs, *segments)))
            if os.path.commonpath([mount_root_abs, resolved_candidate]) != mount_root_abs:
                resolved_candidate = mount_root_abs
            return resolved_candidate, {
                "path_strategy": "mounted_fallback",
                "uri_host": host,
                "mount_root": mount_root_abs,
                "path_sanitized": True,
            }
        resolved = f"//{host}{path_part}".rstrip("/")
        return resolved, {"path_strategy": "posix_unc_like", "uri_host": host, "mount_root": None}

    if os.name == "nt":
        windows_path = path_part.replace("/", "\\")
        if re.match(r"^\\[A-Za-z]:\\", windows_path):
            windows_path = windows_path[1:]
        return windows_path, {"path_strategy": "windows_local_file_uri"}

    return path_part, {"path_strategy": "posix_local_file_uri"}


def file_uri_to_fs_path(uri: str) -> str:
    resolved, _ = resolve_source_path(uri)
    return resolved


def inspect_source_path(source: str, *, expect_directory: bool) -> dict:
    diagnostic = {
        "configured_source": source or "",
        "resolved_path": None,
        "runtime_os": os.name,
        "runtime_platform": sys.platform,
        "path_strategy": None,
        "mount_root": get_file_uri_mount_root(),
        "exists": False,
        "is_directory": False,
        "is_file": False,
        "readable": False,
        "directory_exists": False,
        "error_kind": None,
        "os_error_detail": None,
    }

    if not source:
        diagnostic["error_kind"] = "not_configured"
        return diagnostic

    resolved_source, resolution_meta = resolve_source_path(source)
    diagnostic.update(resolution_meta)
    diagnostic["resolved_path"] = resolved_source

    try:
        diagnostic["exists"] = os.path.exists(resolved_source)
        diagnostic["is_directory"] = os.path.isdir(resolved_source)
        diagnostic["is_file"] = os.path.isfile(resolved_source)
        diagnostic["directory_exists"] = diagnostic["is_directory"]
        if diagnostic["exists"]:
            diagnostic["readable"] = os.access(resolved_source, os.R_OK)
    except OSError as exc:
        diagnostic["error_kind"] = "os_error"
        diagnostic["os_error_detail"] = str(exc)
        return diagnostic

    if not diagnostic["exists"]:
        diagnostic["error_kind"] = "missing"
        return diagnostic
    if not diagnostic["readable"]:
        diagnostic["error_kind"] = "access_denied"
        return diagnostic
    if expect_directory and not diagnostic["is_directory"]:
        diagnostic["error_kind"] = "not_directory"
        return diagnostic
    if not expect_directory and diagnostic["is_directory"]:
        diagnostic["error_kind"] = "is_directory"
        return diagnostic
    if not expect_directory and not diagnostic["is_file"]:
        diagnostic["error_kind"] = "not_file"
        return diagnostic

    return diagnostic


def get_source_label(source: str) -> str:
    if not source:
        return "—"
    if not is_file_uri(source):
        return os.path.basename(source)
    parsed = urlsplit(source)
    name = os.path.basename(unquote(parsed.path.rstrip("/")))
    return name or parsed.netloc or source


def parse_reglement_date(value: str | None) -> date | None:
    if value is None:
        return None
    value = value.strip()
    if not DATE_REGEX.fullmatch(value):
        return None
    try:
        return datetime.strptime(value, "%Y%m%d").date()
    except ValueError:
        return None


def read_text_file(source: str) -> tuple[str | None, str | None]:
    resolved_source = source
    try:
        if is_file_uri(source):
            resolved_source = file_uri_to_fs_path(source)
        else:
            resolved_source = source

        with open(resolved_source, "rb") as f:
            content = f.read()
        try:
            return content.decode("utf-8"), None
        except UnicodeDecodeError:
            return content.decode("latin-1"), None
    except FileNotFoundError:
        return None, f"Fichier introuvable : {resolved_source}"
    except PermissionError:
        return None, f"Accès refusé : {resolved_source}"
    except IsADirectoryError:
        return None, f"Chemin invalide (dossier) : {resolved_source}"
    except OSError as exc:
        return None, f"Impossible de lire le fichier : {resolved_source} ({exc.__class__.__name__})"


def list_reglement_files(directory: str) -> tuple[list[str], list[str]]:
    lookup_dir = file_uri_to_fs_path(directory) if is_file_uri(directory) else directory

    if not os.path.exists(lookup_dir):
        return [], [f"Dossier introuvable : {lookup_dir}"]
    if not os.path.isdir(lookup_dir):
        return [], [f"Chemin invalide (pas un dossier) : {lookup_dir}"]

    try:
        filenames = sorted(
            [f for f in os.listdir(lookup_dir) if is_reglement_text_filename(f)],
            key=lambda f: f.lower(),
        )
    except PermissionError:
        return [], [f"Accès refusé au dossier : {lookup_dir}"]
    except OSError as exc:
        return [], [f"Impossible de lire le dossier : {lookup_dir} ({exc.__class__.__name__})"]

    if is_file_uri(directory):
        base_uri = directory if directory.endswith("/") else f"{directory}/"
        return [f"{base_uri}{quote(f)}" for f in filenames], []
    return [os.path.join(directory, f) for f in filenames], []


def unique_paths(paths: list[str]) -> list[str]:
    ordered: list[str] = []
    seen: set[str] = set()
    for path in paths:
        key = path.lower()
        if key in seen:
            continue
        seen.add(key)
        ordered.append(path)
    return ordered


def _parse_single_file(path: str) -> tuple[str, list[dict], str | None]:
    """Worker for parallel parsing. Returns (path, rows, error)."""
    text, error = read_text_file(path)
    if error:
        return path, [], error
    return path, parse_lines(text or ""), None


def load_rows_from_paths(paths: list[str]) -> tuple[list[dict], list[str], list[str]]:
    """Parallel loader: reads and parses files concurrently in a thread pool."""
    unique = unique_paths(paths)
    if not unique:
        return [], [], []

    rows: list[dict] = []
    source_files: list[str] = []
    warnings: list[str] = []

    # Parallel parsing (huge win for many history files)
    with ThreadPoolExecutor(max_workers=MAX_PARSE_WORKERS) as executor:
        results = list(executor.map(_parse_single_file, unique))

    for path, parsed_rows, error in results:
        if error:
            warnings.append(error)
            continue
        source_files.append(path)
        rows.extend(parsed_rows)

    return rows, source_files, warnings


def parse_article_lines(text: str) -> dict[str, str]:
    articles: dict[str, str] = {}
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        parts = [p.strip() for p in line.split(";")]
        if len(parts) < 2 or not parts[0]:
            continue
        articles[parts[0]] = parts[1] or parts[0]
    return articles


def parse_facture_lines(text: str, article_lookup: dict[str, str] | None = None) -> list[dict]:
    lookup = article_lookup or {}
    results: list[dict] = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        parts = [p.strip() for p in line.split(";")]
        if len(parts) < 15:
            continue

        facture_number = parts[0]
        facture_date = parse_reglement_date(parts[2])
        cam_value = parts[5] or ""
        cam_match = CAM_REGEX.search(cam_value) or CAM_REGEX.search(parts[1])
        cam = cam_match.group(1) if cam_match else None
        site = get_site(cam) if cam else normalize_site(parts[4])
        article_code = parts[7]

        try:
            package_quantity = float((parts[9] or "0").replace(",", "."))
            quantity = float((parts[10] or "0").replace(",", "."))
            amount = float((parts[-2] or "0").replace(",", "."))
        except ValueError:
            continue

        results.append({
            "facture_number": facture_number,
            "reference": parts[1],
            "facture_date": facture_date,
            "facture_date_iso": facture_date.isoformat() if facture_date else None,
            "document_type": parts[3],
            "raw_site": parts[4],
            "site": site,
            "cam": cam,
            "client_code": parts[6],
            "article_code": article_code,
            "article_name": lookup.get(article_code, article_code),
            "unit": parts[8],
            "package_quantity": package_quantity,
            "quantity": quantity,
            "amount": amount,
        })
    return results


def load_article_lookup(path: str) -> tuple[dict[str, str], list[str]]:
    if not path:
        return {}, []
    text, error = read_text_file(path)
    if error:
        return {}, [error]
    return parse_article_lines(text or ""), []


def safe_part(parts: list[str], index: int, default: str = "") -> str:
    if index < 0 or index >= len(parts):
        return default
    value = parts[index]
    return value if value is not None else default


def parse_client_lines(text: str) -> list[dict]:
    """Parse CLIENT.txt lines into client master-data records.

    Example line:
    CLF00045;0960711SNC000;1;DIRCT;EPS;GOVSF;SF383;CT;EP AOUICHAOUI MANNOUBIA;LOCF;
    A01;CHVSFX08;SUPSFX1;CAM38;EP AOUICHAOUI MANNOUBIA;Route Gabes Km 6 - Cité Nasr;
    SFAX;34.702715;10.706005;7;ESP30J;4900;2
    """
    results: list[dict] = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        parts = [p.strip() for p in line.split(";")]
        if len(parts) < 14 or not parts[0]:
            continue

        # NEW: skip clients flagged as inactive (last field == "1")
        status_code = safe_part(parts, len(parts) - 1)
        if status_code == "1":
            continue

        client_code = parts[0]
        cam_value = safe_part(parts, 13)
        cam_match = CAM_REGEX.search(cam_value)
        cam = cam_match.group(1) if cam_match else None

        client_name = safe_part(parts, 14) or safe_part(parts, 8) or client_code
        address = safe_part(parts, 15)
        city = safe_part(parts, 16)
        payment_terms = safe_part(parts, 20)

        results.append({
            "client_code": client_code,
            "tax_id": safe_part(parts, 1),
            "cam": cam,
            "site": get_site(cam) if cam else "Inconnu",
            "client_name": client_name,
            "address": address,
            "city": city,
            "payment_terms": payment_terms,
        })
    return results


def load_client_lookup(path: str) -> tuple[dict[str, dict], dict[str, list[dict]], list[str]]:
    if not path:
        return {}, {}, []
    text, error = read_text_file(path)
    if error:
        return {}, {}, [error]

    records = parse_client_lines(text or "")
    if not records:
        return {}, {}, ["CLIENT.txt : aucune ligne client exploitable."]

    lookup: dict[str, dict] = {}
    by_cam: dict[str, list[dict]] = defaultdict(list)
    for record in records:
        lookup[record["client_code"]] = record
        cam = record.get("cam")
        if cam:
            by_cam[cam].append(record)

    return lookup, dict(by_cam), []


def parse_decimal(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text.replace(" ", "").replace(",", "."))
    except ValueError:
        return None


def build_etatmarge_issue_message(filename: str, message: str) -> str:
    return f"{filename} : {message}"


def parse_etatmarge_rows_with_diagnostics(
    rows: list[list[object]],
) -> tuple[dict[str, float], list[str], list[str]]:
    filename = "etatmarge"
    
    # Check if file has any data
    has_data = any(
        any(str(cell).strip() for cell in row if cell is not None)
        for row in rows
    )
    if not has_data:
        return {}, [build_etatmarge_issue_message(filename, "fichier vide.")], []

    # Pad all rows to same length
    max_cols = max(len(row) for row in rows) if rows else 0
    normalized_rows = []
    for row in rows:
        padded_row = list(row) + [None] * (max_cols - len(row))
        normalized_row = [
            normalize_token(str(cell)) if cell is not None else ""
            for cell in padded_row
        ]
        normalized_rows.append(normalized_row)

    # ← NEW: Find ALL candidate header rows, pick the best one
    header_candidates = []

    for idx in range(min(50, len(normalized_rows))):
        normalized_cells = normalized_rows[idx]
        
        article_candidates = [
            i for i, cell in enumerate(normalized_cells)
            if cell and (
                cell == "article" or
                cell == "art" or
                "codearticle" in cell or
                (cell.startswith("article") and len(cell) <= 12)
            )
        ]
        
        tva_candidates = [
            i for i, cell in enumerate(normalized_cells)
            if cell and (
                cell == "tva" or
                cell == "taxe" or
                "tauxtva" in cell or
                "tauxdetva" in cell or
                cell.endswith("tva")
            )
        ]

        # BOTH must exist in this row, and in DIFFERENT columns
        if article_candidates and tva_candidates:
            if article_candidates[0] != tva_candidates[0]:
                header_candidates.append({
                    'row_index': idx,
                    'article_index': article_candidates[0],
                    'tva_index': tva_candidates[0],
                })

    if not header_candidates:
        # Original error diagnostics (keep as-is)
        article_found_anywhere = False
        tva_found_anywhere = False
        article_rows = []
        tva_rows = []
        
        for idx in range(min(50, len(normalized_rows))):
            row = normalized_rows[idx]
            if any("article" in c for c in row if c):
                article_found_anywhere = True
                article_rows.append(idx + 1)
            if any("tva" in c or "taxe" in c for c in row if c):
                tva_found_anywhere = True
                tva_rows.append(idx + 1)

        if not article_found_anywhere and not tva_found_anywhere:
            return {}, [
                build_etatmarge_issue_message(
                    filename,
                    "ligne d'en-tête introuvable (colonnes Article / Tva non détectées)."
                )
            ], []
        
        if article_found_anywhere and not tva_found_anywhere:
            return {}, [
                build_etatmarge_issue_message(
                    filename,
                    f"colonne Tva manquante (Article trouvé aux lignes {article_rows[:3]})."
                )
            ], []
        
        if tva_found_anywhere and not article_found_anywhere:
            return {}, [
                build_etatmarge_issue_message(
                    filename,
                    f"colonne Article manquante (Tva trouvé aux lignes {tva_rows[:3]})."
                )
            ], []
        
        return {}, [
            build_etatmarge_issue_message(
                filename,
                f"Article et Tva détectés sur des lignes différentes (Article: lignes {article_rows[:3]}, Tva: lignes {tva_rows[:3]})."
            )
        ], []

    # ← NEW: Pick the header with Article in the LEFTMOST position (cleanest layout)
    best_header = min(header_candidates, key=lambda h: h['article_index'])
    header_index = best_header['row_index']
    article_index = best_header['article_index']
    tva_index = best_header['tva_index']

    # Parse data rows after header
    lookup: dict[str, float] = {}
    invalid_tva_articles: list[str] = []

    for row_idx, row in enumerate(rows[header_index + 1:], start=header_index + 2):
        # Ensure row has enough columns
        if len(row) <= max(article_index, tva_index):
            continue

        article_code = normalize_article_code(
            str(row[article_index]) if row[article_index] is not None else ""
        )
        if not article_code:
            continue

        raw_tva = row[tva_index]
        tva_rate = parse_decimal(raw_tva)
        
        if tva_rate is None:
            if str(raw_tva or "").strip():
                invalid_tva_articles.append(f"{article_code} (ligne {row_idx})")
            continue
        
        lookup[article_code] = tva_rate

    # Warnings
    warnings: list[str] = []
    if invalid_tva_articles:
        sample = ", ".join(invalid_tva_articles[:5])
        suffix = "…" if len(invalid_tva_articles) > 5 else ""
        warnings.append(
            build_etatmarge_issue_message(
                filename,
                f"valeur TVA invalide ignorée pour {sample}{suffix}.",
            )
        )

    if not lookup:
        return {}, [
            build_etatmarge_issue_message(
                filename,
                "aucune ligne de données exploitable après l'en-tête."
            )
        ], warnings

    return lookup, [], warnings

@app.get("/api/debug/etatmarge-raw")
async def debug_etatmarge_raw():
    import_context = _cache.get("import_context") or {}
    path = import_context.get("etatmarge_path", "")
    
    if not path or not os.path.isfile(path):
        return JSONResponse({"error": "fichier introuvable", "path": path})
    
    _, ext = os.path.splitext(path)
    ext = ext.casefold()
    result = {
        "path": path,
        "ext": ext,
        "file_size": os.path.getsize(path),
        "reader_used": None,
        "rows": [],
        "error": None,
    }
    
    try:
        if ext in {".xlsx", ".xlsm"}:
            result["reader_used"] = "openpyxl"
            wb = load_workbook(path, read_only=True, data_only=True)
            sheet = wb.active
            result["sheet_name"] = sheet.title
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i >= 20: break
                result["rows"].append([str(c) if c is not None else "" for c in row])
            wb.close()

        elif ext == ".xls":
            result["reader_used"] = "xlrd"
            try:
                import xlrd
                result["xlrd_version"] = xlrd.__VERSION__
                book = xlrd.open_workbook(path)
                result["sheet_names"] = book.sheet_names()
                sheet = book.sheet_by_index(0)
                result["nrows"] = sheet.nrows
                result["ncols"] = sheet.ncols
                for r in range(min(20, sheet.nrows)):
                    result["rows"].append([str(c) for c in sheet.row_values(r)])
            except ImportError:
                result["reader_used"] = "xlrd_NOT_INSTALLED"
                result["error"] = "xlrd non installé!"
            except Exception as e:
                result["reader_used"] = "xlrd_FAILED"
                result["error"] = f"{e.__class__.__name__}: {e}"
                # Try reading as HTML (some .xls are actually HTML)
                try:
                    with open(path, "rb") as f:
                        first_bytes = f.read(200)
                    result["first_bytes_hex"] = first_bytes.hex()
                    result["first_bytes_text"] = first_bytes.decode("latin-1", errors="replace")
                    if b"<html" in first_bytes.lower() or b"<table" in first_bytes.lower():
                        result["file_type_guess"] = "HTML_DISGUISED_AS_XLS"
                    elif first_bytes[:8] == bytes([0xD0,0xCF,0x11,0xE0,0xA1,0xB1,0x1A,0xE1]):
                        result["file_type_guess"] = "REAL_XLS_BINARY (D0CF magic bytes OK)"
                    else:
                        result["file_type_guess"] = "UNKNOWN"
                except Exception as e2:
                    result["bytes_error"] = str(e2)
        else:
            result["reader_used"] = "text_fallback"
            text, err = read_text_file(path)
            result["error"] = err
            result["rows"] = (text or "").splitlines()[:20]

    except Exception as exc:
        result["error"] = f"{exc.__class__.__name__}: {exc}"

    return JSONResponse(result)

def load_etatmarge_lookup(path: str) -> tuple[dict[str, float], list[str]]:
    if not path:
        return {}, []

    filename = os.path.basename(path) or "etatmarge"
    _, ext = os.path.splitext(path)
    ext = ext.casefold()
    try:
        if ext in {".xlsx", ".xlsm"}:
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            workbook.close()
        elif ext == ".xls":
            # ← NEW: Handle legacy .xls files with xlrd
            try:
                import xlrd
                book = xlrd.open_workbook(path)
                sheet = book.sheet_by_index(0)
                rows = [sheet.row_values(r) for r in range(sheet.nrows)]
            except ImportError:
                return {}, [build_etatmarge_issue_message(filename, "xlrd non installé (requis pour .xls).")]
            except Exception as exc:
                return {}, [build_etatmarge_issue_message(filename, f"fichier .xls invalide ({exc.__class__.__name__}).")]
        else:
            # Fallback: text parsing for .txt, .csv
            text, error = read_text_file(path)
            if error:
                return {}, [error]
            rows = []
            for line in (text or "").splitlines():
                raw_line = line.strip()
                if not raw_line:
                    continue
                if "\t" in raw_line:
                    parts = [p.strip() for p in raw_line.split("\t")]
                elif ";" in raw_line:
                    parts = [p.strip() for p in raw_line.split(";")]
                elif "|" in raw_line:
                    parts = [p.strip() for p in raw_line.split("|")]
                else:
                    continue
                rows.append(parts)

        lookup, errors, warnings = parse_etatmarge_rows_with_diagnostics(rows)
        if errors:
            return {}, [build_etatmarge_issue_message(filename, e.split(" : ", 1)[-1]) for e in errors]
        return lookup, [build_etatmarge_issue_message(filename, w.split(" : ", 1)[-1]) for w in warnings]
    except Exception as exc:
        return {}, [build_etatmarge_issue_message(filename, f"fichier invalide ({exc.__class__.__name__}).")]

def list_plain_files(directory: str) -> tuple[list[str], list[str]]:
    lookup_dir = file_uri_to_fs_path(directory) if is_file_uri(directory) else directory

    if not os.path.exists(lookup_dir):
        return [], [f"Dossier introuvable : {lookup_dir}"]
    if not os.path.isdir(lookup_dir):
        return [], [f"Chemin invalide : {lookup_dir}"]

    try:
        filenames = sorted(
            [f for f in os.listdir(lookup_dir)
             if os.path.isfile(os.path.join(lookup_dir, f)) and not f.startswith(".")],
            key=lambda f: f.lower(),
        )
    except PermissionError:
        return [], [f"Accès refusé : {lookup_dir}"]
    except OSError as exc:
        return [], [f"Erreur ({exc.__class__.__name__}) : {lookup_dir}"]

    if is_file_uri(directory):
        base_uri = directory if directory.endswith("/") else f"{directory}/"
        return [f"{base_uri}{quote(f)}" for f in filenames], []
    return [os.path.join(directory, f) for f in filenames], []


def _parse_single_facture_file(args):
    path, article_lookup = args
    text, error = read_text_file(path)
    if error:
        return path, [], error
    return path, parse_facture_lines(text or "", article_lookup), None


def load_facture_lines_from_paths(paths, article_lookup):
    unique = unique_paths(paths)
    if not unique:
        return [], [], []

    rows = []
    source_files = []
    warnings = []

    with ThreadPoolExecutor(max_workers=MAX_PARSE_WORKERS) as executor:
        results = list(executor.map(_parse_single_facture_file, [(p, article_lookup) for p in unique]))

    for path, parsed, error in results:
        if error:
            warnings.append(error)
            continue
        source_files.append(path)
        rows.extend(parsed)

    return rows, source_files, warnings


def get_facture_bounds(rows):
    dates = [r.get("facture_date") for r in rows if isinstance(r.get("facture_date"), date)]
    if not dates:
        return None, None
    return min(dates), max(dates)


def build_big_factures(rows, etatmarge_lookup=None):
    tva_lookup = etatmarge_lookup or {}
    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        facture_number = row.get("facture_number")
        if facture_number:
            grouped[facture_number].append(row)

    big_factures = []
    for facture_number, facture_rows in grouped.items():
        first = facture_rows[0]
        base_total = sum(line.get("amount", 0.0) for line in facture_rows)
        tva_total = 0.0
        for line in facture_rows:
            article_code = normalize_article_code(line.get("article_code"))
            tva_rate = tva_lookup.get(article_code, 0.0)
            if not tva_rate:
                continue
            tva_total += line.get("amount", 0.0) * tva_rate / 100.0
        timbre = 1.0
        total_amount = round(base_total + tva_total + timbre, 3)
        total_quantity = round(sum(line.get("quantity", 0.0) for line in facture_rows), 3)
        big_factures.append({
            "facture_number": facture_number,
            "facture_date": first.get("facture_date"),
            "facture_date_iso": first.get("facture_date_iso"),
            "site": first.get("site") or get_site(first.get("cam")),
            "cam": first.get("cam"),
            "client_code": first.get("client_code"),
            "reference": first.get("reference"),
            "line_count": len(facture_rows),
            "total_amount": total_amount,
            "base_total": round(base_total, 3),
            "tva_total": round(tva_total, 3),
            "timbre": timbre,
            "total_quantity": total_quantity,
            "articles_count": len({line.get("article_code") for line in facture_rows if line.get("article_code")}),
            "lines": facture_rows,
        })

    return sorted(
        big_factures,
        key=lambda item: (item.get("facture_date") or date.min, item.get("facture_number") or ""),
        reverse=True,
    )


def filter_big_factures_by_date(rows, start_date, end_date):
    filtered = []
    for row in rows:
        facture_date = row.get("facture_date")
        if facture_date is None:
            continue
        if start_date is not None and facture_date < start_date:
            continue
        if end_date is not None and facture_date > end_date:
            continue
        filtered.append(row)
    return filtered


def build_cam_facture_payload(cam, factures, *, mode, source_files=None, date_range=None, warnings=None):
    cam_factures = [f for f in factures if f.get("cam") == cam]
    cam_factures.sort(
        key=lambda item: (item.get("facture_date") or date.min, item.get("facture_number") or ""),
        reverse=True,
    )

    article_acc: dict[str, dict] = defaultdict(
        lambda: {"article_name": "", "quantity": 0.0, "amount": 0.0, "line_count": 0}
    )
    for facture in cam_factures:
        for line in facture.get("lines", []):
            article_code = line.get("article_code") or "—"
            article = article_acc[article_code]
            article["article_name"] = line.get("article_name") or article_code
            article["quantity"] += line.get("quantity", 0.0)
            article["amount"] += line.get("amount", 0.0)
            article["line_count"] += 1

    top_articles = sorted(
        [
            {
                "article_code": code,
                "article_name": v["article_name"],
                "quantity": round(v["quantity"], 3),
                "amount": round(v["amount"], 3),
                "line_count": v["line_count"],
            }
            for code, v in article_acc.items()
        ],
        key=lambda item: (item["amount"], item["quantity"]),
        reverse=True,
    )

    factures_payload = [
        {
            "facture_number": f["facture_number"],
            "facture_date_iso": f["facture_date_iso"],
            "client_code": f["client_code"],
            "reference": f["reference"],
            "line_count": f["line_count"],
            "articles_count": f["articles_count"],
            "total_quantity": f["total_quantity"],
            "total_amount": f["total_amount"],
            "top_articles_preview": [
                line.get("article_name") or line.get("article_code")
                for line in sorted(f.get("lines", []), key=lambda item: item.get("amount", 0.0), reverse=True)[:3]
            ],
        }
        for f in cam_factures
    ]

    return {
        "cam": cam,
        "site": get_site(cam),
        "mode": mode,
        "date_range": date_range,
        "source_files": source_files or [],
        "warnings": warnings or [],
        "nb_factures": len(cam_factures),
        "total_vente": round(sum(f["total_amount"] for f in cam_factures), 3),
        "top_articles": top_articles[:8],
        "factures": factures_payload,
    }


def filter_rows_by_date(rows, start_date, end_date):
    filtered = []
    for row in rows:
        reglement_date = row.get("reglement_date")
        if reglement_date is None:
            continue
        if start_date is not None and reglement_date < start_date:
            continue
        if end_date is not None and reglement_date > end_date:
            continue
        filtered.append(row)
    return filtered


def parse_iso_date(value: str) -> date:
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError("Date invalide.") from exc


def parse_lines(text: str):
    """Optimized parser - pre-compiled regex, minimal allocations."""
    results = []
    type_map_keys = list(TYPE_MAP.keys())
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        parts = line.split(";")
        if len(parts) < 11:
            continue

        code = parts[0]
        prefix = None
        for candidate in type_map_keys:
            if code.startswith(candidate):
                prefix = candidate
                break
        if prefix is None:
            continue

        try:
            amount = float(parts[-1].replace(",", "."))
        except ValueError:
            continue

        ref2 = parts[1]
        cam_match = CAM_REGEX.search(ref2)
        cam = cam_match.group(1) if cam_match else None
        reglement_date = parse_reglement_date(parts[2])

        results.append({
            "code": code,
            "cam": cam,
            "site": normalize_site(parts[4]),
            "type_key": prefix,
            "type_label": TYPE_MAP[prefix],
            "amount": amount,
            "reglement_date": reglement_date,
            "reglement_date_iso": reglement_date.isoformat() if reglement_date else None,
        })
    return results


@app.get("/", response_class=HTMLResponse)
async def index():
    async with aiofiles.open("static/index.html", encoding="utf-8") as f:
        return await f.read()


def date_to_iso(value):
    return value.isoformat() if isinstance(value, date) else None


def get_rows_bounds(rows):
    dates = [r.get("reglement_date") for r in rows if isinstance(r.get("reglement_date"), date)]
    if not dates:
        return None, None
    return min(dates), max(dates)


def _invalidate_derived_caches():
    """Clear precomputed payloads when raw data changes."""
    _cache["_default_payload"] = None
    _cache["_status_payload"] = None
    _cache["_range_cache"] = {}


def reload_cache() -> None:
    import_context = _cache.get("import_context") or {}
    using_uploaded_folder = bool(import_context.get("active"))

    current_path = (import_context.get("current_path") if using_uploaded_folder else DEFAULT_CURRENT_REGLEMENT_FILE) or ""
    history_dir = (import_context.get("history_path") if using_uploaded_folder else DEFAULT_HISTORY_REGLEMENTS_DIR) or ""
    article_path = (import_context.get("article_path") if using_uploaded_folder else "") or ""
    etatmarge_path = (import_context.get("etatmarge_path") if using_uploaded_folder else "") or ""
    facture_path = (import_context.get("facture_path") if using_uploaded_folder else "") or ""
    factures_dir = (import_context.get("factures_path") if using_uploaded_folder else "") or ""
    client_path = (import_context.get("client_path") if using_uploaded_folder else "") or ""
    warnings: list[str] = []
    if using_uploaded_folder:
        warnings.extend(import_context.get("warnings", []))

    current_diagnostic = inspect_source_path(current_path, expect_directory=False)
    history_diagnostic = inspect_source_path(history_dir, expect_directory=True)
    sync_info = {
        "required": False,
        "mode": "uploaded_folder" if using_uploaded_folder else "configured_paths",
        "status": "ready",
        "message": "Chargement depuis le dossier importé." if using_uploaded_folder else "Chargement depuis les chemins configurés.",
        "cache_root": import_context.get("root_path") if using_uploaded_folder else None,
        "local_current_path": current_path if using_uploaded_folder else None,
        "local_history_path": history_dir if using_uploaded_folder else None,
        "copied_history_files": 0,
    }

    if not current_path and not history_dir:
        if using_uploaded_folder:
            warnings.append("Le dossier importé doit contenir REGLEMENT.txt et le sous-dossier Réglements.")
        else:
            warnings.append("Aucune donnée importée. Importez le dossier racine Fichiers Sources.")
        _cache.update({
            "all_rows": [], "current_rows": [], "source_files": [], "current_source_files": [],
            "article_lookup": {}, "etatmarge_lookup": {}, "etatmarge_warnings": [],
            "all_facture_lines": [], "all_big_factures": [], "current_big_factures": [],
            "facture_source_files": [], "current_facture_source_files": [],
            "client_lookup": {}, "clients_by_cam": {}, "client_warnings": [],
            "warnings": warnings, "current_warnings": [], "loaded_at": time.time(),
            "coverage_start": None, "coverage_end": None, "history_file_count": 0,
            "facture_coverage_start": None, "facture_coverage_end": None,
            "facture_history_file_count": 0,
            "source_diagnostics": {"current": current_diagnostic, "history": history_diagnostic},
            "sync": sync_info,
        })
        _invalidate_derived_caches()
        return

    history_files: list[str] = []
    if history_dir:
        history_files, dir_warnings = list_reglement_files(history_dir)
        warnings.extend(dir_warnings)

    # Parallel loading: history + current + article + etatmarge + factures all at once
    history_rows: list[dict] = []
    history_sources: list[str] = []
    if history_files:
        history_rows, history_sources, hist_warnings = load_rows_from_paths(history_files)
        warnings.extend(hist_warnings)

    current_rows: list[dict] = []
    current_sources: list[str] = []
    cur_warnings: list[str] = []
    if current_path:
        current_rows, current_sources, cur_warnings = load_rows_from_paths([current_path])
        warnings.extend(cur_warnings)

    all_rows = history_rows + current_rows
    all_sources = history_sources + current_sources
    coverage_start, coverage_end = get_rows_bounds(all_rows)
    article_lookup, article_warnings = load_article_lookup(article_path)
    warnings.extend(article_warnings)
    etatmarge_lookup, etatmarge_warnings = load_etatmarge_lookup(etatmarge_path)
    warnings.extend(etatmarge_warnings)
    client_lookup, clients_by_cam, client_warnings = load_client_lookup(client_path)
    warnings.extend(client_warnings)

    facture_history_files: list[str] = []
    if factures_dir:
        facture_history_files, facture_dir_warnings = list_plain_files(factures_dir)
        warnings.extend(facture_dir_warnings)

    facture_current_rows: list[dict] = []
    facture_current_sources: list[str] = []
    if facture_path:
        facture_current_rows, facture_current_sources, fcw = load_facture_lines_from_paths([facture_path], article_lookup)
        warnings.extend(fcw)

    facture_history_rows: list[dict] = []
    facture_history_sources: list[str] = []
    if facture_history_files:
        facture_history_rows, facture_history_sources, fhw = load_facture_lines_from_paths(facture_history_files, article_lookup)
        warnings.extend(fhw)

    all_facture_lines = facture_history_rows + facture_current_rows
    current_big_factures = build_big_factures(facture_current_rows, etatmarge_lookup)
    all_big_factures = build_big_factures(all_facture_lines, etatmarge_lookup)
    facture_coverage_start, facture_coverage_end = get_facture_bounds(all_facture_lines)

    _cache.update({
        "all_rows": all_rows,
        "current_rows": current_rows,
        "source_files": all_sources,
        "current_source_files": current_sources,
        "article_lookup": article_lookup,
        "etatmarge_lookup": etatmarge_lookup,
        "etatmarge_warnings": etatmarge_warnings,
        "all_facture_lines": all_facture_lines,
        "all_big_factures": all_big_factures,
        "current_big_factures": current_big_factures,
        "facture_source_files": facture_history_sources + facture_current_sources,
        "current_facture_source_files": facture_current_sources,
        "client_lookup": client_lookup,
        "clients_by_cam": clients_by_cam,
        "client_warnings": client_warnings,
        "warnings": warnings,
        "current_warnings": cur_warnings if current_path else [],
        "loaded_at": time.time(),
        "coverage_start": date_to_iso(coverage_start),
        "coverage_end": date_to_iso(coverage_end),
        "history_file_count": len(history_files),
        "facture_coverage_start": date_to_iso(facture_coverage_start),
        "facture_coverage_end": date_to_iso(facture_coverage_end),
        "facture_history_file_count": len(facture_history_files),
        "source_diagnostics": {"current": current_diagnostic, "history": history_diagnostic},
        "sync": sync_info,
    })
    _invalidate_derived_caches()


def get_or_reload_cache() -> dict:
    if _cache["loaded_at"] is None:
        import_context = _cache.get("import_context") or {}
        if import_context.get("status") == "processing":
            return _cache
        reload_cache()
    return _cache


def get_source_status() -> dict:
    if _cache.get("_status_payload") is not None:
        return _cache["_status_payload"]

    cache = _cache
    loaded_at = cache["loaded_at"]
    import_context = cache.get("import_context", {})
    import_status = import_context.get("status") or ("ready" if loaded_at is not None else "idle")
    using_uploaded_folder = bool(import_context.get("active"))
    current_source = import_context.get("current_path") if using_uploaded_folder else DEFAULT_CURRENT_REGLEMENT_FILE
    history_source = import_context.get("history_path") if using_uploaded_folder else DEFAULT_HISTORY_REGLEMENTS_DIR
    article_source = import_context.get("article_path") if using_uploaded_folder else ""
    etatmarge_source = import_context.get("etatmarge_path") if using_uploaded_folder else ""
    facture_source = import_context.get("facture_path") if using_uploaded_folder else ""
    factures_source = import_context.get("factures_path") if using_uploaded_folder else ""
    client_source = import_context.get("client_path") if using_uploaded_folder else ""

    status = {
        "loaded": loaded_at is not None,
        "loaded_at": loaded_at,
        "coverage_start": cache["coverage_start"],
        "coverage_end": cache["coverage_end"],
        "source_file_count": len(cache["source_files"]),
        "history_file_count": cache["history_file_count"] if loaded_at is not None else int(import_context.get("history_file_count") or 0),
        "facture_source_file_count": len(cache.get("facture_source_files", [])),
        "facture_history_file_count": cache.get("facture_history_file_count", 0) if loaded_at is not None else int(import_context.get("facture_history_file_count") or 0),
        "has_data": bool(cache["all_rows"]),
        "has_facture_data": bool(cache.get("all_big_factures")),
        "warnings": cache["warnings"],
        "current_source_label": get_source_label(current_source) if current_source else "—",
        "history_source_label": get_source_label(history_source) if history_source else "—",
        "article_source_label": get_source_label(article_source) if article_source else "—",
        "etatmarge_source_label": get_source_label(etatmarge_source) if etatmarge_source else "—",
        "facture_source_label": get_source_label(facture_source) if facture_source else "—",
        "factures_source_label": get_source_label(factures_source) if factures_source else "—",
        "client_source_label": get_source_label(client_source) if client_source else "—",
        "source_mode": "uploaded_folder" if using_uploaded_folder else "configured_paths",
        "import_status": import_status,
        "import_error_message": import_context.get("error_message"),
        "import_context": {
            "active": bool(import_context.get("active")),
            "status": import_status,
            "uploaded_at": import_context.get("uploaded_at"),
            "error_message": import_context.get("error_message"),
            "root_name": import_context.get("root_name"),
            "root_path": import_context.get("root_path"),
        },
        "uploaded_root_name": import_context.get("root_name"),
        "uploaded_root_path": import_context.get("root_path"),
        "expected_current_name": "REGLEMENT.txt",
        "expected_history_name": "Réglements",
        "expected_article_name": "ARTICLE",
        "expected_etatmarge_name": "etatmarge",
        "expected_facture_name": "FACTURE",
        "expected_factures_name": "Factures",
        "expected_client_name": "CLIENT",
        "current_found": import_context.get("current_found", bool(current_source)),
        "history_found": import_context.get("history_found", bool(history_source)),
        "article_found": import_context.get("article_found", bool(article_source)),
        "etatmarge_found": import_context.get("etatmarge_found", bool(etatmarge_source)),
        "facture_found": import_context.get("facture_found", bool(facture_source)),
        "factures_found": import_context.get("factures_found", bool(factures_source)),
        "client_found": import_context.get("client_found", bool(client_source)),
        "facture_coverage_start": cache.get("facture_coverage_start"),
        "facture_coverage_end": cache.get("facture_coverage_end"),
        "etatmarge_warnings": cache.get("etatmarge_warnings", []),
        "etatmarge_lookup_size": len(cache.get("etatmarge_lookup", {})),
        "client_warnings": cache.get("client_warnings", []),
        "client_lookup_size": len(cache.get("client_lookup", {})),
        "runtime_label": f"{os.name}/{sys.platform}",
        "current_diagnostic": cache.get("source_diagnostics", {}).get("current", {}),
        "history_diagnostic": cache.get("source_diagnostics", {}).get("history", {}),
        "sync_required": cache.get("sync", {}).get("required", False),
        "sync_mode": cache.get("sync", {}).get("mode", "direct_read"),
        "sync_status": cache.get("sync", {}).get("status", "not_required"),
        "sync_message": cache.get("sync", {}).get("message"),
        "local_cache_root": cache.get("sync", {}).get("cache_root"),
        "local_current_path": cache.get("sync", {}).get("local_current_path"),
        "local_history_path": cache.get("sync", {}).get("local_history_path"),
        "copied_history_files": cache.get("sync", {}).get("copied_history_files", 0),
    }
    _cache["_status_payload"] = status
    return status


def build_import_results(status: dict) -> list[dict]:
    etatmarge_warnings = status.get("etatmarge_warnings", [])
    etatmarge_warning_msg = " ".join(etatmarge_warnings) if etatmarge_warnings else ""
    etatmarge_lookup_size = int(status.get("etatmarge_lookup_size") or 0)
    client_warnings = status.get("client_warnings", [])
    client_warning_msg = " ".join(client_warnings) if client_warnings else ""
    client_lookup_size = int(status.get("client_lookup_size") or 0)
    return [
        {"key": "current_reglement", "label": "REGLEMENT.txt", "required": True,
         "status": "ok" if status.get("current_found") else "missing",
         "kind": "ok" if status.get("current_found") else "err",
         "file": status.get("current_source_label") if status.get("current_found") else None,
         "message": "Fichier chargé." if status.get("current_found") else "Fichier requis introuvable."},
        {"key": "history_reglements", "label": "Réglements (historique)", "required": True,
         "status": "ok" if status.get("history_found") else "missing",
         "kind": "ok" if status.get("history_found") else "err",
         "file": status.get("history_source_label") if status.get("history_found") else None,
         "message": f"{status.get('history_file_count', 0)} fichier(s) chargé(s)." if status.get("history_found") else "Dossier requis introuvable."},
        {"key": "article", "label": "ARTICLE", "required": True,
         "status": "ok" if status.get("article_found") else "missing",
         "kind": "ok" if status.get("article_found") else "err",
         "file": status.get("article_source_label") if status.get("article_found") else None,
         "message": "Fichier chargé." if status.get("article_found") else "Fichier requis introuvable."},
        {"key": "etatmarge", "label": "etatmarge", "required": False,
         "status": "optional" if not status.get("etatmarge_found") else ("error" if etatmarge_warnings and etatmarge_lookup_size == 0 else ("warning" if etatmarge_warnings else "ok")),
         "kind": "warn" if not status.get("etatmarge_found") else ("err" if etatmarge_warnings and etatmarge_lookup_size == 0 else ("warn" if etatmarge_warnings else "ok")),
         "file": status.get("etatmarge_source_label") if status.get("etatmarge_found") else None,
         "message": "Fichier optionnel non fourni." if not status.get("etatmarge_found") else (etatmarge_warning_msg or "Fichier chargé.")},
        {"key": "facture", "label": "FACTURE", "required": True,
         "status": "ok" if status.get("facture_found") else "missing",
         "kind": "ok" if status.get("facture_found") else "err",
         "file": status.get("facture_source_label") if status.get("facture_found") else None,
         "message": "Fichier chargé." if status.get("facture_found") else "Fichier requis introuvable."},
        {"key": "history_factures", "label": "Factures (historique)", "required": True,
         "status": "ok" if status.get("factures_found") else "missing",
         "kind": "ok" if status.get("factures_found") else "err",
         "file": status.get("factures_source_label") if status.get("factures_found") else None,
         "message": f"{status.get('facture_history_file_count', 0)} fichier(s) chargé(s)." if status.get("factures_found") else "Dossier requis introuvable."},
        {"key": "client", "label": "CLIENT", "required": False,
         "status": "optional" if not status.get("client_found") else ("error" if client_warnings and client_lookup_size == 0 else ("warning" if client_warnings else "ok")),
         "kind": "warn" if not status.get("client_found") else ("err" if client_warnings and client_lookup_size == 0 else ("warn" if client_warnings else "ok")),
         "file": status.get("client_source_label") if status.get("client_found") else None,
         "message": "Fichier optionnel non fourni (analyse clients inactifs désactivée)." if not status.get("client_found") else (client_warning_msg or f"{client_lookup_size} client(s) chargé(s).")},
    ]

@app.get("/api/site/{site_code}/comparison")
async def get_site_comparison(
    site_code: str,
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
):
    site = normalize_site(site_code)
    cache = get_or_reload_cache()

    if start_date and end_date:
        try:
            start = parse_iso_date(start_date)
            end = parse_iso_date(end_date)
        except ValueError:
            return JSONResponse({"detail": "Date invalide."}, status_code=400)
    else:
        start, end = get_rows_bounds(cache["current_rows"])
        if not start or not end:
            return {"site": site, "comparison": None}

    return {"site": site, "comparison": build_comparison_payload_scoped(cache, start, end, site=site)}

@app.on_event("startup")
async def startup() -> None:
    reload_cache()


def build_upload_payload(rows, filename, *, mode="default", source_files=None, date_range=None, warnings=None):
    cam_data: dict[str, dict] = defaultdict(
        lambda: {"total": 0.0, "count": 0, "by_type": defaultdict(lambda: {"amount": 0.0, "count": 0})}
    )
    cam_rows_count = 0

    for r in rows:
        cam = r["cam"]
        if cam is not None:
            d = cam_data[cam]
            d["total"] += r["amount"]
            d["count"] += 1
            t = d["by_type"][r["type_label"]]
            t["amount"] += r["amount"]
            t["count"] += 1
            cam_rows_count += 1

    ranked = sorted(
        [
            {
                "cam": cam,
                "site": get_site(cam),
                "total": round(v["total"], 3),
                "total_count": v["count"],
                "esp": round(v["by_type"].get("Espèces", {}).get("amount", 0.0), 3),
                "trt": round(v["by_type"].get("Traite", {}).get("amount", 0.0), 3),
                "chq": round(v["by_type"].get("Chèque", {}).get("amount", 0.0), 3),
                "esp_count": v["by_type"].get("Espèces", {}).get("count", 0),
                "trt_count": v["by_type"].get("Traite", {}).get("count", 0),
                "chq_count": v["by_type"].get("Chèque", {}).get("count", 0),
                "by_type": {k: {"amount": round(vt["amount"], 3), "count": vt["count"]} for k, vt in v["by_type"].items()},
            }
            for cam, v in cam_data.items()
        ],
        key=lambda x: x["total"],
        reverse=True,
    )

    for i, item in enumerate(ranked, 1):
        item["rank"] = i

    all_types_summary = defaultdict(lambda: {"amount": 0.0, "count": 0})
    for r in rows:
        all_types_summary[r["type_label"]]["amount"] += r["amount"]
        all_types_summary[r["type_label"]]["count"] += 1

    sites_acc: dict[str, dict] = defaultdict(
        lambda: {"amount": 0.0, "count": 0, "cams": set(),
                 "esp_count": 0, "trt_count": 0, "chq_count": 0,
                 "esp_amount": 0.0, "trt_amount": 0.0, "chq_amount": 0.0}
    )
    for r in rows:
        cam = r["cam"]
        site = get_site(cam) if cam is not None else r["site"]
        sites_acc[site]["amount"] += r["amount"]
        sites_acc[site]["count"] += 1
        if cam is not None:
            sites_acc[site]["cams"].add(cam)
        label = r["type_label"]
        if label == "Espèces":
            sites_acc[site]["esp_count"] += 1
            sites_acc[site]["esp_amount"] += r["amount"]
        elif label == "Traite":
            sites_acc[site]["trt_count"] += 1
            sites_acc[site]["trt_amount"] += r["amount"]
        elif label == "Chèque":
            sites_acc[site]["chq_count"] += 1
            sites_acc[site]["chq_amount"] += r["amount"]

    sites_summary = {
        k: {
            "amount": round(v["amount"], 3),
            "count": v["count"],
            "cam_count": len(v["cams"]),
            "esp_count": v["esp_count"],
            "trt_count": v["trt_count"],
            "chq_count": v["chq_count"],
            "esp_amount": round(v["esp_amount"], 3),
            "trt_amount": round(v["trt_amount"], 3),
            "chq_amount": round(v["chq_amount"], 3),
        }
        for k, v in sorted(sites_acc.items())
    }

    return {
        "rows": ranked,
        "grand_total": round(sum(r["amount"] for r in rows), 3),
        "grand_count": len(rows),
        "lines_parsed": len(rows),
        "active_cams": len(cam_data),
        "skipped_rows": len([r for r in rows if r["cam"] is None and r["site"] == "Inconnu"]),
        "rows_without_cam": len([r for r in rows if r["cam"] is None]),
        "rows_without_cam_with_site": len([r for r in rows if r["cam"] is None and r["site"] != "Inconnu"]),
        "rows_with_cam": cam_rows_count,
        "types_summary": {k: {"amount": round(v["amount"], 3), "count": v["count"]} for k, v in all_types_summary.items()},
        "sites_summary": sites_summary,
        "filename": filename,
        "source_label": filename,
        "mode": mode,
        "source_files": source_files or [],
        "date_range": date_range,
        "warnings": warnings or [],
    }


# ── API endpoints ─────────────────────────────────────────────────────────────

@app.get("/api/default")
async def get_default_dashboard():
    cache = get_or_reload_cache()

    # Return precomputed payload if available
    if cache.get("_default_payload") is not None:
        return cache["_default_payload"]

    import_context = cache.get("import_context", {})
    using_uploaded_folder = bool(import_context.get("active"))
    current_source = import_context.get("current_path") if using_uploaded_folder else DEFAULT_CURRENT_REGLEMENT_FILE
    label = f"Mois en cours · {get_source_label(current_source)}"
    default_warnings = cache["current_warnings"] if cache["current_source_files"] else cache["warnings"]

    payload = build_upload_payload(
        cache["current_rows"],
        label,
        mode="default",
        source_files=cache["current_source_files"],
        warnings=default_warnings,
    )
    cur_start, cur_end = get_rows_bounds(cache["current_rows"])
    if cur_start and cur_end:
        payload["comparison"] = build_comparison_payload_scoped(cache, cur_start, cur_end)
    else:
        payload["comparison"] = None
        payload["facture_totals"] = {
            "nb_factures": len(cache.get("current_big_factures", [])),
            "total_ventes": round(sum(f["total_amount"] for f in cache.get("current_big_factures", [])), 3),
        }
    _cache["_default_payload"] = payload
    return payload


@app.get("/api/range")
async def get_dashboard_for_range(
    start_date: str = Query(..., description="Date de début au format AAAA-MM-JJ"),
    end_date: str = Query(..., description="Date de fin au format AAAA-MM-JJ"),
):
    try:
        start = parse_iso_date(start_date)
        end = parse_iso_date(end_date)
    except ValueError:
        return JSONResponse({"detail": "Date invalide. Format attendu AAAA-MM-JJ."}, status_code=400)

    if start > end:
        return JSONResponse({"detail": "La date de début doit être antérieure ou égale à la date de fin."}, status_code=400)

    # Cache lookup
    cache_key = f"{start.isoformat()}::{end.isoformat()}"
    range_cache = _cache.get("_range_cache") or {}
    if cache_key in range_cache:
        return range_cache[cache_key]

    label = f"Période du {start.strftime('%d/%m/%Y')} au {end.strftime('%d/%m/%Y')}"
    cache = get_or_reload_cache()
    filtered_rows = filter_rows_by_date(cache["all_rows"], start, end)

    payload = build_upload_payload(
        filtered_rows,
        label,
        mode="date_range",
        source_files=cache["source_files"],
        date_range={"start": start.isoformat(), "end": end.isoformat()},
        warnings=cache["warnings"],
    )

    # Cache with LRU-style trim (max 20 entries)
    if len(range_cache) >= 20:
        # Drop oldest entry
        try:
            oldest = next(iter(range_cache))
            del range_cache[oldest]
        except StopIteration:
            pass
    payload["comparison"] = build_comparison_payload_scoped(cache, start, end)
    range_cache[cache_key] = payload
    _cache["_range_cache"] = range_cache
    return payload


@app.get("/api/filter")
async def get_dashboard_for_filter(
    start_date: str = Query(..., alias="start"),
    end_date: str = Query(..., alias="end"),
):
    return await get_dashboard_for_range(start_date=start_date, end_date=end_date)


@app.get("/api/cam/{cam_code}")
async def get_cam_facture_detail(
    cam_code: str,
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
):
    cam = cam_code.strip().upper()
    cache = get_or_reload_cache()
    warnings = list(cache["warnings"])

    if (start_date and not end_date) or (end_date and not start_date):
        return JSONResponse({"detail": "Renseignez les deux dates pour filtrer les factures CAM."}, status_code=400)

    if start_date and end_date:
        try:
            start = parse_iso_date(start_date)
            end = parse_iso_date(end_date)
        except ValueError:
            return JSONResponse({"detail": "Date invalide. Format attendu AAAA-MM-JJ."}, status_code=400)
        if start > end:
            return JSONResponse({"detail": "La date de début doit être antérieure ou égale à la date de fin."}, status_code=400)
        factures = filter_big_factures_by_date(cache.get("all_big_factures", []), start, end)
        mode = "date_range"
        source_files = cache.get("facture_source_files", [])
        date_range = {"start": start.isoformat(), "end": end.isoformat()}
    else:
        factures = cache.get("current_big_factures", [])
        mode = "default"
        source_files = cache.get("current_facture_source_files", [])
        date_range = None

    if not cache.get("all_big_factures"):
        warnings.append("Aucune donnée facture. Importez ARTICLE, FACTURE et le dossier Factures.")
    if start_date and end_date:
        comparison = build_comparison_payload_scoped(cache, start, end, cam=cam)
    else:
        def_start, def_end = get_rows_bounds(cache["current_rows"])
        comparison = build_comparison_payload_scoped(cache, def_start, def_end, cam=cam) if def_start and def_end else None

    result = build_cam_facture_payload(cam, factures, mode=mode, source_files=source_files, date_range=date_range, warnings=warnings)
    result["comparison"] = comparison
    return result


@app.get("/api/cam/{cam_code}/inactive-clients")
async def get_cam_inactive_clients(
    cam_code: str,
    start_date: str = Query(..., description="Date de début au format AAAA-MM-JJ"),
    end_date: str = Query(..., description="Date de fin au format AAAA-MM-JJ"),
):
    """Clients rattachés à cette CAM (via CLIENT.txt) qui n'ont AUCUNE facture
    dans la période [start_date, end_date]. Pour chacun, renvoie la date et le
    numéro de sa dernière facture connue (toutes périodes confondues)."""
    cam = cam_code.strip().upper()

    try:
        start = parse_iso_date(start_date)
        end = parse_iso_date(end_date)
    except ValueError:
        return JSONResponse({"detail": "Date invalide. Format attendu AAAA-MM-JJ."}, status_code=400)
    if start > end:
        return JSONResponse({"detail": "La date de début doit être antérieure ou égale à la date de fin."}, status_code=400)

    cache = get_or_reload_cache()
    warnings: list[str] = []

    clients_by_cam = cache.get("clients_by_cam") or {}
    cam_clients = clients_by_cam.get(cam, [])

    if not cache.get("client_lookup"):
        warnings.append("CLIENT.txt introuvable ou vide. Importez CLIENT.txt pour activer l'analyse des clients inactifs.")

    if not cam_clients:
        return {
            "cam": cam,
            "site": get_site(cam),
            "date_range": {"start": start.isoformat(), "end": end.isoformat()},
            "total_clients": 0,
            "active_clients": 0,
            "inactive_count": 0,
            "inactive_clients": [],
            "warnings": warnings,
        }

    all_facture_lines = cache.get("all_facture_lines") or []

    # Single pass over every facture line: figure out (a) which client codes
    # billed something inside the requested range, and (b) each client's most
    # recent facture ever, regardless of range.
    active_codes: set[str] = set()
    last_facture_by_client: dict[str, dict] = {}

    for line in all_facture_lines:
        code = line.get("client_code")
        fd = line.get("facture_date")
        if not code or fd is None:
            continue

        if start <= fd <= end:
            active_codes.add(code)

        current_best = last_facture_by_client.get(code)
        if current_best is None or fd > current_best["date"]:
            last_facture_by_client[code] = {
                "date": fd,
                "facture_number": line.get("facture_number"),
            }

    inactive_clients = []
    for client in cam_clients:
        code = client.get("client_code")
        if not code or code in active_codes:
            continue

        last = last_facture_by_client.get(code)
        inactive_clients.append({
            "client_code": code,
            "client_name": client.get("client_name") or code,
            "city": client.get("city") or "",
            "address": client.get("address") or "",
            "payment_terms": client.get("payment_terms") or "",
            "last_facture_date": last["date"].isoformat() if last else None,
            "last_facture_number": last["facture_number"] if last else None,
            "days_since_last": (end - last["date"]).days if last else None,
            "never_invoiced": last is None,
        })

    # Never-invoiced clients first (most urgent), then oldest last-purchase first.
    inactive_clients.sort(key=lambda c: (
        0 if c["never_invoiced"] else 1,
        c["last_facture_date"] or "",
    ))

    return {
        "cam": cam,
        "site": get_site(cam),
        "date_range": {"start": start.isoformat(), "end": end.isoformat()},
        "total_clients": len(cam_clients),
        "active_clients": len(cam_clients) - len(inactive_clients),
        "inactive_count": len(inactive_clients),
        "inactive_clients": inactive_clients,
        "warnings": warnings,
    }


@app.get("/api/status")
async def source_status():
    return get_source_status()


async def process_import_async(import_root: str) -> None:
    loop = asyncio.get_running_loop()
    try:
        await loop.run_in_executor(IMPORT_EXECUTOR, reload_cache)
        import_context = _cache.get("import_context") or {}
        if import_context.get("active") and import_context.get("root_path") == import_root:
            import_context["status"] = "ready"
            import_context.pop("error_message", None)
    except Exception as exc:
        logger.exception("Import background processing failed")
        import_context = _cache.get("import_context") or {}
        if import_context.get("active") and import_context.get("root_path") == import_root:
            import_context["status"] = "error"
            import_context["error_message"] = str(exc)


@app.post("/api/refresh")
async def refresh_data():
    reload_cache()
    return get_source_status()


async def _save_upload_file(upload: UploadFile, destination_path: str, semaphore: asyncio.Semaphore) -> int:
    """Save an uploaded file with concurrency control."""
    async with semaphore:
        file_size = 0
        os.makedirs(os.path.dirname(destination_path), exist_ok=True)
        async with aiofiles.open(destination_path, "wb") as out:
            while True:
                chunk = await upload.read(CHUNK_SIZE)
                if not chunk:
                    break
                file_size += len(chunk)
                if file_size > MAX_SINGLE_FILE_SIZE:
                    raise ValueError(
                        f"Fichier trop volumineux ({format_size(file_size)} > {format_size(MAX_SINGLE_FILE_SIZE)})."
                    )
                await out.write(chunk)
        await upload.close()
        return file_size


@app.post("/api/import-folder")
async def import_folder(files: list[UploadFile] = File(...)):
    if not files:
        return JSONResponse(
            {"detail": "Aucun fichier reçu. Sélectionnez le dossier Fichiers Sources."},
            status_code=400,
        )

    # Quick size pre-check
    total_size = sum(upload.size or 0 for upload in files)
    if total_size > MAX_TOTAL_UPLOAD_SIZE:
        return JSONResponse(
            {
                "detail": (
                    f"La taille totale ({format_size(total_size)}) dépasse la limite "
                    f"autorisée ({format_size(MAX_TOTAL_UPLOAD_SIZE)})."
                ),
                "error": {
                    "code": "UPLOAD_SIZE_EXCEEDED",
                    "total_size": total_size,
                    "max_size": MAX_TOTAL_UPLOAD_SIZE,
                    "file_count": len(files),
                },
            },
            status_code=413,
        )

    import_root = tempfile.mkdtemp(prefix="sind_reglement_import_")
    uploaded_root_name: str | None = None
    warnings: list[str] = []
    upload_results: list[dict] = []
    detected_history_file_count = 0
    detected_facture_history_file_count = 0

    try:
        # First pass: validate & prepare tasks
        upload_tasks = []
        semaphore = asyncio.Semaphore(MAX_CONCURRENT_UPLOADS)

        for upload in files:
            filename_raw = upload.filename or "fichier_inconnu"
            raw_rel_path = filename_raw.replace("\\", "/").strip("/")

            if not raw_rel_path:
                upload_results.append({"file": filename_raw, "kind": "err", "message": "Nom de fichier vide."})
                await upload.close()
                continue

            segments = [s for s in raw_rel_path.split("/") if s not in {"", ".", ".."}]
            if not segments:
                upload_results.append({"file": raw_rel_path, "kind": "err", "message": "Chemin invalide."})
                await upload.close()
                continue

            if len(segments) > 1:
                uploaded_root_name = uploaded_root_name or segments[0]
                relative_segments = segments[1:]
            else:
                relative_segments = segments

            if not relative_segments:
                await upload.close()
                continue

            filename = relative_segments[-1]
            parent_segments = relative_segments[:-1]
            is_root_level = len(parent_segments) == 0
            in_reglements_dir = any(is_reglements_dir_name(seg) for seg in parent_segments)
            in_factures_dir = any(is_factures_dir_name(seg) for seg in parent_segments)
            is_monthly_file = filename.casefold() == "reglement.txt"
            is_history_file = in_reglements_dir and is_reglement_text_filename(filename)
            is_article_file = is_root_level and is_named_source_file(filename, "ARTICLE")
            is_etatmarge_file = is_root_level and is_named_excel_source_file(filename, "etatmarge")
            is_current_facture_file = is_root_level and is_named_source_file(filename, "FACTURE")
            is_history_facture_file = in_factures_dir
            is_client_file = is_root_level and is_named_source_file(filename, "CLIENT")

            if not (is_monthly_file or is_history_file or is_article_file or is_etatmarge_file
                    or is_current_facture_file or is_history_facture_file or is_client_file):
                upload_results.append({"file": raw_rel_path, "kind": "warn", "message": "Ignoré (hors périmètre)."})
                await upload.close()
                continue

            destination_path = os.path.join(import_root, *relative_segments)
            if is_history_file:
                detected_history_file_count += 1
            if is_history_facture_file:
                detected_facture_history_file_count += 1

            upload_tasks.append((upload, destination_path, raw_rel_path))

        # Parallel upload with concurrency control
        async def save_and_track(upload, dest, rel):
            try:
                await _save_upload_file(upload, dest, semaphore)
                return {"file": rel, "kind": "ok", "message": "Fichier reçu."}
            except Exception as e:
                return {"file": rel, "kind": "err", "message": f"{e}"}

        save_results = await asyncio.gather(
            *[save_and_track(u, d, r) for u, d, r in upload_tasks],
            return_exceptions=False,
        )
        upload_results.extend(save_results)
        saved_files = sum(1 for r in save_results if r.get("kind") == "ok")

        if saved_files == 0:
            shutil.rmtree(import_root, ignore_errors=True)
            return JSONResponse(
                {
                    "detail": "Le dossier importé est vide ou invalide.",
                    "error": {"code": "EMPTY_IMPORT", "message": "Aucun fichier exploitable reçu."},
                    "upload_file_results": upload_results,
                },
                status_code=400,
            )

        # Cleanup previous import
        previous_context = _cache.get("import_context") or {}
        previous_root = previous_context.get("root_path")
        if previous_root and previous_root != import_root:
            shutil.rmtree(previous_root, ignore_errors=True)

        discovered = discover_uploaded_sources(import_root)
        expected_root_name = "Fichiers Sources"
        if uploaded_root_name:
            discovered["root_name"] = uploaded_root_name
            if normalize_token(uploaded_root_name) != normalize_token(expected_root_name):
                warnings.append(f"Dossier racine: {uploaded_root_name}. Attendu: {expected_root_name}.")
        else:
            discovered["root_name"] = expected_root_name

        discovered["warnings"] = warnings + discovered.get("warnings", [])
        discovered["history_file_count"] = detected_history_file_count
        discovered["facture_history_file_count"] = detected_facture_history_file_count

        _cache.update({
            "all_rows": [], "current_rows": [], "source_files": [], "current_source_files": [],
            "article_lookup": {}, "etatmarge_lookup": {}, "etatmarge_warnings": [],
            "all_facture_lines": [], "all_big_factures": [], "current_big_factures": [],
            "facture_source_files": [], "current_facture_source_files": [],
            "client_lookup": {}, "clients_by_cam": {}, "client_warnings": [],
            "warnings": [], "current_warnings": [], "loaded_at": None,
            "coverage_start": None, "coverage_end": None, "history_file_count": 0,
            "facture_coverage_start": None, "facture_coverage_end": None,
            "facture_history_file_count": 0, "source_diagnostics": {}, "sync": {},
        })
        _cache["import_context"] = {**discovered, "active": True, "uploaded_at": time.time()}
        _invalidate_derived_caches()

        # Trigger parse in background thread — do it synchronously here so the
        # response reflects the loaded data (matches original behavior)
        loop = asyncio.get_running_loop()
        await loop.run_in_executor(IMPORT_EXECUTOR, reload_cache)

        payload = get_source_status()
        payload["import_results"] = build_import_results(payload)
        payload["upload_file_results"] = upload_results
        payload["uploaded_files"] = saved_files
        return payload
    except Exception as exc:
        shutil.rmtree(import_root, ignore_errors=True)
        logger.exception("Import failed")
        return JSONResponse(
            {
                "detail": "Import interrompu. Vérifiez les fichiers et réessayez.",
                "error": {
                    "code": "IMPORT_FAILED",
                    "message": f"{exc.__class__.__name__}: {exc}",
                },
                "upload_file_results": upload_results,
            },
            status_code=500,
        )